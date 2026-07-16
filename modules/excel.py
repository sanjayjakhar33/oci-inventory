"""Workbook export helpers using openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from modules.utils import safe_value, sanitize_sheet_name

logger = logging.getLogger(__name__)


def create_workbook() -> Workbook:
    """Return a workbook with an empty default sheet removed."""
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


class ExcelWorkbookBuilder:
    """Create the inventory workbook with a standardized worksheet layout."""

    def __init__(self, output_dir: str = "output") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_workbook(
        self,
        workbook_rows: dict[str, list[dict[str, object]]],
        compartment_name: str,
        compartment_ocid: str,
    ) -> Path:
        """Build the workbook and return its output path."""
        safe_name = "".join(ch if ch.isalnum() else "_" for ch in compartment_name.strip() or compartment_ocid)
        workbook_path = self.output_dir / f"OCI_Inventory_{safe_name}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        ws.append(["Metric", "Value"])
        ws.append(["Compartment Name", compartment_name])
        ws.append(["Compartment OCID", compartment_ocid])
        for module_name, rows in workbook_rows.items():
            ws.append([f"{module_name} Row Count", len(rows)])

        for sheet_name, rows in workbook_rows.items():
            sheet = wb.create_sheet(title=sanitize_sheet_name(sheet_name))
            if rows:
                headers = list(rows[0].keys())
                sheet.append(headers)
                for row in rows:
                    sheet.append([safe_value(row.get(header, "")) for header in headers])
            self._format_sheet(sheet)

        self._format_sheet(ws)
        wb.save(workbook_path)
        logger.info("Workbook written to %s", workbook_path)
        return workbook_path

    def _format_sheet(self, worksheet) -> None:
        """Apply workbook formatting: freeze panes, bold headers, autofilter, autosize."""
        if worksheet.max_row == 0:
            return
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
        worksheet.auto_filter.ref = worksheet.dimensions
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    cell.value = ""
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)


def export_to_excel(data: list[dict[str, Any]], output_path: str) -> str:
    """Backwards-compatible wrapper for single-sheet export."""
    wb = Workbook()
    ws = wb.active
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([safe_value(row.get(header, "")) for header in headers])
    wb.save(output_path)
    return output_path
